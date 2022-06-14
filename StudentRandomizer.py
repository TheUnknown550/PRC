import random as rn
from openpyxl import Workbook, load_workbook


#ใส่ชื่อคนห้องสายวิทย์   ***Edit!!***
#SciM5 = ['ใบบัว','ทีเจ','ภูผา','ต้นน้ำ','ภูมิ','จั๊ม','ดิน','มอส','พูห์','แซนเดียร์','โปรด','บีม','ชมพู่','กาย','ฟลุ๊คฟลิ๊ก','พราว','เนต','เฟม','นล','ดรณ์','ทิม','หมี','มาร์ค','วา','เอ','วารี','บี','เอิร์ท','ต่างรัฐ']
#SciM4 = ['น้ำปาย','พิมพ์ชนก','กันตพัฒน์','ชยาบดินทร์','ชินดนัย','พชร','ติณณ์','พัชรพรรณ','ธีรภัทร','โชติวัฒนา','ชีวัธนัย','ธัญญาดา','ศรสวรรค์','ศิวกรณ์','พิมพ์ลดา','รวิสรา','ปภังกร','รอยล','รอยล','เลนนี','เกวลิน','สิรวิชญ์','อติคุณ','ธราเทพ','ชัชพล','ชนันท์ภัสส์','ปัณณธร','วีระกรานต์','ธรรศ','ธีรสิทธิ์','ธีรดนย์','สุธีมนต์','ชนิดาภา','ณปก','ทิชาพร','ธนกมล','พัชสุดา','อัณณา','วรทัต','วริษา','อรุณรัชช์','รินรดา','พัสกร','ตรอง','วรัทยา','ณัฐณิชา','นัทธมน']
SciM5 = ['60',"61","62",'63','64','65','66','67','68','69','70',"71","72",'73','74','75','76','77','78','79','80',"81","82",'83','84','85','86','87','88','89','90',"91","92",'93','94','95','96','97','98','99','100',"101","102",'103','104','105']
SciM4 = ["1","2",'3','4','5','6','7','8','9','10',"11","12",'13','14','15','16','17','18','19','20',"21","22",'23','24','25','26','27','28','29','30',"31","32",'33','34','35','36','37','38','39','40',"41","42",'43','44','45','46','47','48','49','50',"51","52",'53','54','55','56','57','58','59',]
Scipair5 = []
Scipair4 = []

#ใส่ชื่อคนห้องสานศิล   ***Edit!!***
ArtM4 = ['Matt','FookFick','Hana','Able','Chris']
ArtM5 = ['Bibua','Yok','beebell','Kan','Mos','John','Sam']
Artpair5 = []
Artpair4 = []

#คนที่ได้ 2 คน Sci   ***Edit!!***
Get1 = 'ภูมิ'
Get2 = 'ต้นน้ำ'
Get3 = 'ภูผา'
Get4 = 'ทีเจ'
Get5 = 'ใบบัว'

#คู่ในสาย Art   ***Edit!!***
Pair1 = 'Mos'
Pairs1 = 'Yok'

Pair2 = 'Sam'
Pairs2 = 'John'

Pair3 = ''
Pairs3 = ''

Pair4 = ''
Pairs4 = ''

Pair5 = ''
Pairs5 = ''


def Sci():
    for i in range(len(SciM5)):
        rand = rn.randint(0,len(SciM4)-1)

        Scipair4.append(SciM4[rand])
        Scipair5.append(SciM5[i])
        SciM4.remove(SciM4[rand])

def Art():

    #Remove Team   ***Edit!!(Remove Comment)***
    ArtM5.remove(Pairs1)
    ArtM5.remove(Pairs2)
    #ArtM5.remove(Pair3)
    #ArtM5.remove(Pair4)
    #ArtM5.remove(Pair5)

    for i in range(len(ArtM4)):
        rand = rn.randint(0,len(ArtM4)-1)

        Artpair5.append(ArtM5[i])
        Artpair4.append(ArtM4[rand])
        ArtM4.remove(ArtM4[rand])


if __name__ == "__main__":
    #Access Excel
    wb = Workbook()
    ws = wb.active

    #Randomizer
    Sci()
    Art()

    #Sci
    ws['A2'].value = "SCI"
    ws['A3'].value = "M5"
    ws['B3'].value = "M4"
    for i in range(len(Scipair5)):
        ws['A'+str(i+4)].value = Scipair5[i]
        ws['B'+str(i+4)].value = Scipair4[i]
        #Change Group Team
        if Scipair5[i] == Get1 or Scipair5[i] == Get2 or Scipair5[i] == Get3 or Scipair5[i] == Get4 or Scipair5[i] == Get5:
            rand = rn.randint(0,len(SciM4)-1)
            ws['C'+str(i+4)].value = SciM4[i]

    #Art
    ws['A'+str(i+7)].value = 'Art'
    ws['A'+str(i+8)].value = "M5"
    ws['C'+str(i+8)].value = "M4"
    for n in range(len(Artpair5)):
        ws['A'+str(i+9+n)].value = Artpair5[n]
        ws['C'+str(i+9+n)].value = Artpair4[n]
        
        #Change Group team
        if Artpair5[n] == Pair1:
            ws['B'+str(i+9+n)].value = Pairs1

        if Artpair5[n] == Pair2:
            ws['B'+str(i+9+n)].value = Pairs2

        if Artpair5[n] == Pair3:
            ws['B'+str(i+9+n)].value = Pairs3

        if Artpair5[n] == Pair4:
            ws['B'+str(i+9+n)].value = Pairs4

        if Artpair5[n] == Pair5:
            ws['B'+str(i+9+n)].value = Pairs5
            
    wb.save('C:/Users/mattc/OneDrive/Documents/PRC/StudentPair.xlsx')
